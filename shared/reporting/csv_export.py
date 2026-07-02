"""M21 Reporting Module — CSV and Excel (xlsx) export.

Produces trade-by-trade CSV (stdlib) and Excel workbook (openpyxl) in memory.
Neither function writes to disk; callers receive raw bytes for storage or email.
"""

from __future__ import annotations

import csv
import io

from shared.reporting.models import DailyReport, TradeRecord

_TRADE_HEADERS = [
    "trade_id",
    "symbol",
    "exchange",
    "direction",
    "entry_time_ms",
    "exit_time_ms",
    "entry_price",
    "exit_price",
    "quantity",
    "pnl",
    "pnl_pct",
    "strategy_tag",
    "slippage_pct",
    "markout_1m_pct",
    "markout_5m_pct",
]


def _trade_row(t: TradeRecord) -> list[object]:
    return [
        t.trade_id,
        t.symbol,
        t.exchange,
        t.direction,
        t.entry_time_ms,
        t.exit_time_ms if t.exit_time_ms is not None else "",
        t.entry_price,
        t.exit_price if t.exit_price is not None else "",
        t.quantity,
        t.pnl,
        t.pnl_pct,
        t.strategy_tag,
        t.slippage_pct if t.slippage_pct is not None else "",
        t.markout_1m_pct if t.markout_1m_pct is not None else "",
        t.markout_5m_pct if t.markout_5m_pct is not None else "",
    ]


def trades_to_csv(trades: list[TradeRecord]) -> bytes:
    """Serialise a list of trades to UTF-8 CSV bytes.

    Args:
        trades: Trade records to export.

    Returns:
        UTF-8 encoded CSV bytes with header row.
    """
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_TRADE_HEADERS)
    for t in trades:
        writer.writerow(_trade_row(t))
    return buf.getvalue().encode("utf-8")


def trades_to_excel(
    trades: list[TradeRecord],
    daily_report: DailyReport | None = None,
) -> bytes:
    """Generate an Excel workbook with a Trades sheet and optional Summary sheet.

    Args:
        trades: Trade records for the Trades sheet.
        daily_report: When provided, a Summary sheet is added with key metrics.

    Returns:
        Raw xlsx bytes (openpyxl Workbook serialised to memory).
    """
    import openpyxl  # lazy import; openpyxl optional at module load
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()

    # --- Trades sheet ---
    ws_trades = wb.active
    ws_trades.title = "Trades"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2980B9")

    ws_trades.append(_TRADE_HEADERS)
    for cell in ws_trades[1]:
        cell.font = header_font
        cell.fill = header_fill

    for t in trades:
        ws_trades.append(_trade_row(t))

    for col in ws_trades.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in col),
            default=10,
        )
        ws_trades.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    # --- Summary sheet ---
    if daily_report is not None:
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["Metric", "Value"])
        for cell in ws_summary[1]:
            cell.font = header_font
            cell.fill = header_fill

        def _fmt_opt(v: float | None) -> str:
            return f"{v:.4f}" if v is not None else "N/A"

        rows: list[tuple[str, object]] = [
            ("Date", str(daily_report.date)),
            ("Starting Capital", daily_report.starting_capital),
            ("Total P&L", daily_report.total_pnl),
            ("P&L %", daily_report.total_pnl_pct),
            ("Total Trades", daily_report.total_trades),
            ("Win Rate %", daily_report.win_rate_pct),
            ("Winning Trades", daily_report.winning_trades),
            ("Losing Trades", daily_report.losing_trades),
            ("Sharpe", _fmt_opt(daily_report.sharpe)),
            ("Sortino", _fmt_opt(daily_report.sortino)),
            ("Max Drawdown %", daily_report.max_drawdown_pct),
            ("Avg Slippage %", _fmt_opt(daily_report.avg_slippage_pct)),
        ]
        for label, value in rows:
            ws_summary.append([label, value])

        ws_summary.column_dimensions["A"].width = 22
        ws_summary.column_dimensions["B"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
