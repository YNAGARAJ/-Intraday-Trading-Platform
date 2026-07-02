"""M21 Reporting Module — 20 VERIFY scenarios.

Run via: ``python -m shared.reporting``
"""

from __future__ import annotations

import csv as _csv
import io
from datetime import date

import structlog

from shared.reporting.csv_export import trades_to_csv, trades_to_excel
from shared.reporting.html_report import build_daily_html, build_monthly_html
from shared.reporting.metrics import (
    compute_daily_report,
    compute_max_drawdown,
    compute_monthly_report,
    compute_sharpe,
    compute_sortino,
)
from shared.reporting.models import DailyReport, TradeRecord
from shared.reporting.pdf_report import build_daily_pdf, build_monthly_pdf

logger = structlog.get_logger(__name__)

# ---------------------------------------------------------------------------
# Shared test fixtures
# ---------------------------------------------------------------------------


def _make_trade(
    trade_id: str = "TRD001",
    symbol: str = "RELIANCE",
    direction: str = "LONG",
    pnl: float = 500.0,
    pnl_pct: float = 0.5,
    slippage_pct: float | None = 0.02,
    markout_1m: float | None = 0.3,
    markout_5m: float | None = 0.5,
) -> TradeRecord:
    return TradeRecord(
        trade_id=trade_id,
        symbol=symbol,
        exchange="NSE",
        direction=direction,
        entry_time_ms=1_700_000_000_000,
        exit_time_ms=1_700_003_600_000,
        entry_price=2900.0,
        exit_price=2914.5,
        quantity=100,
        pnl=pnl,
        pnl_pct=pnl_pct,
        strategy_tag="STRAT001",
        slippage_pct=slippage_pct,
        markout_1m_pct=markout_1m,
        markout_5m_pct=markout_5m,
    )


def _two_trades() -> list[TradeRecord]:
    return [
        _make_trade("T1", pnl=500.0, pnl_pct=0.5),
        _make_trade("T2", pnl=-200.0, pnl_pct=-0.2),
    ]


def _daily(
    trades: list[TradeRecord] | None = None,
    capital: float = 1_000_000.0,
    report_date: date = date(2026, 7, 1),
) -> DailyReport:
    return compute_daily_report(report_date, trades or _two_trades(), capital)


# ---------------------------------------------------------------------------
# Scenarios
# ---------------------------------------------------------------------------


def _s01_trade_record_fields() -> bool:
    """S01: TradeRecord stores all fields correctly."""
    t = _make_trade()
    return t.symbol == "RELIANCE" and t.direction == "LONG" and t.pnl == 500.0


def _s02_daily_report_empty_trades() -> bool:
    """S02: compute_daily_report with zero trades → zeros, no Sharpe/Sortino."""
    r = compute_daily_report(date(2026, 7, 1), [], 1_000_000.0)
    return (
        r.total_pnl == 0.0
        and r.total_trades == 0
        and r.sharpe is None
        and r.sortino is None
    )


def _s03_daily_report_pnl_sum() -> bool:
    """S03: compute_daily_report sums trade PnLs correctly."""
    r = _daily()
    return abs(r.total_pnl - 300.0) < 1e-6 and r.total_trades == 2


def _s04_win_rate_calculation() -> bool:
    """S04: Win rate = winning_trades / total_trades * 100."""
    r = _daily()
    return (
        r.winning_trades == 1
        and r.losing_trades == 1
        and abs(r.win_rate_pct - 50.0) < 1e-6
    )


def _s05_sharpe_two_data_points() -> bool:
    """S05: compute_sharpe with 2 returns produces a finite value."""
    s = compute_sharpe([0.01, -0.005])
    return s is not None and abs(s) < 1000.0


def _s06_sharpe_none_for_single_return() -> bool:
    """S06: compute_sharpe returns None with fewer than 2 data points."""
    return compute_sharpe([0.01]) is None and compute_sharpe([]) is None


def _s07_sharpe_none_for_zero_variance() -> bool:
    """S07: compute_sharpe returns None when all returns are identical."""
    return compute_sharpe([0.01, 0.01, 0.01]) is None


def _s08_sortino_none_for_no_losses() -> bool:
    """S08: compute_sortino returns None when no negative returns."""
    return compute_sortino([0.01, 0.02, 0.03]) is None


def _s09_sortino_finite_with_losses() -> bool:
    """S09: compute_sortino returns finite value when negative returns exist."""
    s = compute_sortino([0.01, -0.005, 0.008, -0.003])
    return s is not None and abs(s) < 1000.0


def _s10_max_drawdown_zero_on_all_gains() -> bool:
    """S10: compute_max_drawdown = 0.0 when returns are all non-negative."""
    dd = compute_max_drawdown([0.01, 0.02, 0.0, 0.01])
    return dd == 0.0


def _s11_max_drawdown_computed_correctly() -> bool:
    """S11: max_drawdown detects peak-to-trough correctly."""
    # Wealth: 1 → 1.1 → 0.99 → 1.0 (drawdown at step 2: (1.1-0.99)/1.1 ≈ 10%)
    dd = compute_max_drawdown([0.1, -0.1, 0.0101])
    return 0.09 < dd < 0.12


def _s12_monthly_report_aggregation() -> bool:
    """S12: compute_monthly_report aggregates daily P&Ls correctly."""
    d1 = _daily(report_date=date(2026, 7, 1))
    d2 = _daily(report_date=date(2026, 7, 2))
    m = compute_monthly_report(2026, 7, [d1, d2])
    return (
        abs(m.total_pnl - (d1.total_pnl + d2.total_pnl)) < 1e-6
        and m.total_trades == 4
    )


def _s13_pdf_daily_returns_bytes() -> bool:
    """S13: build_daily_pdf returns non-empty PDF bytes."""
    r = _daily()
    data = build_daily_pdf(r)
    return isinstance(data, bytes) and data[:5] == b"%PDF-"


def _s14_pdf_monthly_returns_bytes() -> bool:
    """S14: build_monthly_pdf returns non-empty PDF bytes."""
    d = _daily()
    m = compute_monthly_report(2026, 7, [d])
    data = build_monthly_pdf(m)
    return isinstance(data, bytes) and data[:5] == b"%PDF-"


def _s15_html_daily_contains_metrics() -> bool:
    """S15: build_daily_html contains Sharpe, Sortino, and trade symbol."""
    r = _daily()
    h = build_daily_html(r)
    return "Sharpe" in h and "Sortino" in h and "RELIANCE" in h


def _s16_html_daily_valid_document() -> bool:
    """S16: build_daily_html returns a valid HTML document."""
    h = build_daily_html(_daily())
    return h.startswith("<!DOCTYPE html>") and "</html>" in h


def _s17_html_monthly_contains_daily_breakdown() -> bool:
    """S17: build_monthly_html includes the daily breakdown table."""
    d = _daily()
    m = compute_monthly_report(2026, 7, [d])
    h = build_monthly_html(m)
    return "Daily Breakdown" in h and "2026-07-01" in h


def _s18_csv_export_header_and_rows() -> bool:
    """S18: trades_to_csv has correct header and one row per trade."""
    trades = _two_trades()
    data = trades_to_csv(trades)
    text = data.decode("utf-8")
    rows = list(_csv.reader(io.StringIO(text)))
    return (
        rows[0][0] == "trade_id"
        and "symbol" in rows[0]
        and len(rows) == len(trades) + 1  # header + trades
    )


def _s19_excel_export_sheets() -> bool:
    """S19: trades_to_excel produces xlsx bytes with Trades and Summary sheets."""
    import openpyxl

    r = _daily()
    data = trades_to_excel(r.trades, daily_report=r)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    return "Trades" in wb.sheetnames and "Summary" in wb.sheetnames


def _s20_markout_curve_in_daily_report() -> bool:
    """S20: Markout curve points are computed from trade markout fields."""
    r = _daily()
    labels = {pt.offset_label for pt in r.markout_curve}
    return "T+1m" in labels and "T+5m" in labels and all(
        pt.sample_count > 0 for pt in r.markout_curve
    )


# ---------------------------------------------------------------------------
# Runner
# ---------------------------------------------------------------------------

_SCENARIOS = [
    _s01_trade_record_fields,
    _s02_daily_report_empty_trades,
    _s03_daily_report_pnl_sum,
    _s04_win_rate_calculation,
    _s05_sharpe_two_data_points,
    _s06_sharpe_none_for_single_return,
    _s07_sharpe_none_for_zero_variance,
    _s08_sortino_none_for_no_losses,
    _s09_sortino_finite_with_losses,
    _s10_max_drawdown_zero_on_all_gains,
    _s11_max_drawdown_computed_correctly,
    _s12_monthly_report_aggregation,
    _s13_pdf_daily_returns_bytes,
    _s14_pdf_monthly_returns_bytes,
    _s15_html_daily_contains_metrics,
    _s16_html_daily_valid_document,
    _s17_html_monthly_contains_daily_breakdown,
    _s18_csv_export_header_and_rows,
    _s19_excel_export_sheets,
    _s20_markout_curve_in_daily_report,
]


def run_verify() -> bool:
    """Execute all 20 VERIFY scenarios. Returns True if all pass."""
    passed = 0
    failed = 0
    for fn in _SCENARIOS:
        label = fn.__name__
        doc = (fn.__doc__ or "").strip()
        try:
            ok = fn()
        except Exception as exc:
            ok = False
            logger.error("verify_scenario_exception", scenario=label, error=str(exc))
        if ok:
            passed += 1
            logger.info("verify_pass", scenario=label, description=doc)
        else:
            failed += 1
            logger.error("verify_fail", scenario=label, description=doc)
    logger.info("VERIFY_SUMMARY", passed=passed, failed=failed, total=len(_SCENARIOS))
    return failed == 0
