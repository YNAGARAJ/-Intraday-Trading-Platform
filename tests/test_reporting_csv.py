"""Tests for M21 CSV and Excel export."""

from __future__ import annotations

import csv
import io
from datetime import date

import pytest

from shared.reporting.csv_export import trades_to_csv, trades_to_excel
from shared.reporting.metrics import compute_daily_report
from shared.reporting.models import TradeRecord


def _trade(
    tid: str = "T1",
    pnl: float = 500.0,
    pnl_pct: float = 0.5,
    slippage_pct: float | None = 0.02,
    markout_1m: float | None = 0.3,
    markout_5m: float | None = 0.5,
) -> TradeRecord:
    return TradeRecord(
        trade_id=tid,
        symbol="RELIANCE",
        exchange="NSE",
        direction="LONG",
        entry_time_ms=1_700_000_000_000,
        exit_time_ms=1_700_003_600_000,
        entry_price=2900.0,
        exit_price=2914.5,
        quantity=100,
        pnl=pnl,
        pnl_pct=pnl_pct,
        strategy_tag="STRAT001",
        slippage_pct=slippage_pct,
        markout_1m_pct=markout_1m,
        markout_5m_pct=markout_5m,
    )


class TestTradesToCsv:
    def test_returns_bytes(self) -> None:
        assert isinstance(trades_to_csv([_trade()]), bytes)

    def test_utf8_decodable(self) -> None:
        data = trades_to_csv([_trade()])
        text = data.decode("utf-8")
        assert len(text) > 0

    def test_header_row_present(self) -> None:
        data = trades_to_csv([_trade()])
        rows = list(csv.reader(io.StringIO(data.decode("utf-8"))))
        assert rows[0][0] == "trade_id"

    def test_expected_columns_in_header(self) -> None:
        data = trades_to_csv([_trade()])
        rows = list(csv.reader(io.StringIO(data.decode("utf-8"))))
        header = rows[0]
        for col in ("symbol", "exchange", "direction", "pnl", "strategy_tag"):
            assert col in header

    def test_one_data_row_per_trade(self) -> None:
        trades = [_trade(tid="T1"), _trade(tid="T2"), _trade(tid="T3")]
        data = trades_to_csv(trades)
        rows = list(csv.reader(io.StringIO(data.decode("utf-8"))))
        assert len(rows) == 4  # header + 3 trades

    def test_empty_trades_returns_header_only(self) -> None:
        data = trades_to_csv([])
        rows = list(csv.reader(io.StringIO(data.decode("utf-8"))))
        assert len(rows) == 1  # header only

    def test_trade_id_in_row(self) -> None:
        data = trades_to_csv([_trade(tid="XYZ-999")])
        assert b"XYZ-999" in data

    def test_symbol_in_row(self) -> None:
        t = TradeRecord(
            trade_id="T1",
            symbol="INFY",
            exchange="NSE",
            direction="LONG",
            entry_time_ms=1_700_000_000_000,
            exit_time_ms=None,
            entry_price=1500.0,
            exit_price=None,
            quantity=50,
            pnl=0.0,
            pnl_pct=0.0,
            strategy_tag="STRAT002",
        )
        data = trades_to_csv([t])
        assert b"INFY" in data

    def test_pnl_value_in_row(self) -> None:
        data = trades_to_csv([_trade(pnl=1234.56)])
        assert b"1234.56" in data

    def test_none_fields_export_as_empty(self) -> None:
        t = _trade(slippage_pct=None, markout_1m=None, markout_5m=None)
        data = trades_to_csv([t])
        rows = list(csv.reader(io.StringIO(data.decode("utf-8"))))
        header = rows[0]
        row = rows[1]
        row_dict = dict(zip(header, row, strict=True))
        assert row_dict.get("slippage_pct", "") == ""

    def test_multiple_trades_correct_order(self) -> None:
        trades = [_trade(tid="A"), _trade(tid="B"), _trade(tid="C")]
        data = trades_to_csv(trades)
        rows = list(csv.reader(io.StringIO(data.decode("utf-8"))))
        assert rows[1][0] == "A"
        assert rows[2][0] == "B"
        assert rows[3][0] == "C"


class TestTradesToExcel:
    def test_returns_bytes(self) -> None:
        assert isinstance(trades_to_excel([_trade()]), bytes)

    def test_valid_xlsx_bytes(self) -> None:
        import openpyxl

        data = trades_to_excel([_trade()])
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert wb is not None

    def test_trades_sheet_exists(self) -> None:
        import openpyxl

        data = trades_to_excel([_trade()])
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "Trades" in wb.sheetnames

    def test_summary_sheet_requires_daily_report(self) -> None:
        import openpyxl

        r = compute_daily_report(date(2026, 7, 1), [_trade()], 1_000_000.0)
        data = trades_to_excel([_trade()], daily_report=r)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "Summary" in wb.sheetnames

    def test_no_summary_sheet_without_daily_report(self) -> None:
        import openpyxl

        data = trades_to_excel([_trade()])
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "Summary" not in wb.sheetnames

    def test_trades_sheet_has_header_row(self) -> None:
        import openpyxl

        data = trades_to_excel([_trade()])
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Trades"]
        headers = [c.value for c in ws[1]]
        assert "trade_id" in headers

    def test_trades_sheet_data_rows(self) -> None:
        import openpyxl

        trades = [_trade(tid="T1"), _trade(tid="T2")]
        data = trades_to_excel(trades)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Trades"]
        assert ws.max_row == 3  # header + 2 data rows

    def test_empty_trades_produces_valid_xlsx(self) -> None:
        import openpyxl

        data = trades_to_excel([])
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "Trades" in wb.sheetnames

    def test_summary_sheet_has_sharpe_row(self) -> None:
        import openpyxl

        r = compute_daily_report(
            date(2026, 7, 1),
            [_trade(tid="A"), _trade(pnl=-100.0, pnl_pct=-0.1, tid="B")],
            1_000_000.0,
        )
        data = trades_to_excel(r.trades, daily_report=r)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Summary"]
        labels = [str(ws.cell(row=i, column=1).value) for i in range(1, ws.max_row + 1)]
        assert any("Sharpe" in (lbl or "") for lbl in labels)

    @pytest.mark.parametrize("n_trades", [0, 1, 5, 20])
    def test_row_count_matches_trades(self, n_trades: int) -> None:
        import openpyxl

        trades = [_trade(tid=f"T{i}") for i in range(n_trades)]
        data = trades_to_excel(trades)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Trades"]
        assert ws.max_row == n_trades + 1  # header row
